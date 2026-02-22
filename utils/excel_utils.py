import gc
import logging
import os
from datetime import date as dt_date

from openpyxl import Workbook, load_workbook

from domain.import_policy import ImportPolicy
from domain.records import MandatoryExpenseRecord, Record
from domain.reports import Report
from utils.csv_utils import (
    _parse_transfer_row,
    _restore_missing_transfers,
    _validate_transfer_integrity,
)
from utils.import_core import (
    ImportSummary,
    norm_key,
    parse_import_row,
    record_type_name,
    safe_type,
)

logger = logging.getLogger(__name__)

DATA_HEADERS = [
    "date",
    "type",
    "wallet_id",
    "category",
    "amount_original",
    "currency",
    "rate_at_operation",
    "amount_kzt",
    "description",
    "period",
    "transfer_id",
    "from_wallet_id",
    "to_wallet_id",
]
MANDATORY_HEADERS = [
    "date",
    "type",
    "category",
    "amount_original",
    "currency",
    "rate_at_operation",
    "amount_kzt",
    "description",
    "period",
]


def _safe_str(value):
    return "" if value is None else str(value)


def _resolve_get_rate(currency_service):
    if currency_service is None:
        from app.services import CurrencyService

        currency_service = CurrencyService()
    return currency_service.get_rate


def report_to_xlsx(report: Report, filepath: str) -> None:
    """Export report view (fixed amounts) to XLSX. Read-only format."""
    wb = Workbook()
    ws = wb.active
    if ws is not None:
        ws.title = "Report"
        ws.append([report.statement_title, "", "", ""])
        ws.append(["Date", "Type", "Category", "Amount (KZT)"])
        ws.append(["", "", "", "Fixed amounts by operation-time FX rates"])

    if (getattr(report, "initial_balance", 0) != 0 or report.is_opening_balance) and ws is not None:
        ws.append(["", report.balance_label, "", f"{report.initial_balance:.2f}"])

    for record in sorted(
        report.records(),
        key=lambda r: (0, r.date) if isinstance(r.date, dt_date) else (1, dt_date.max),
    ):
        typ = record_type_name(record)
        if typ == "income":
            record_type = "Income"
        elif typ == "mandatory_expense":
            record_type = "Mandatory Expense"
        else:
            record_type = "Expense"
        if ws is not None:
            record_date = (
                record.date.isoformat() if isinstance(record.date, dt_date) else record.date
            )
            ws.append([record_date, record_type, record.category, f"{record.amount_kzt:.2f}"])

    total = report.total_fixed()
    records_total = sum(r.signed_amount_kzt() for r in report.records())
    if ws is not None:
        ws.append(["SUBTOTAL", "", "", f"{records_total:.2f}"])
        ws.append(["FINAL BALANCE", "", "", f"{total:.2f}"])

    summary_year, monthly_rows = report.monthly_income_expense_rows()
    summary_ws = wb.create_sheet("Yearly Report")
    if summary_ws is not None:
        summary_ws.append([f"Month ({summary_year})", "Income (KZT)", "Expense (KZT)"])
        total_income = 0.0
        total_expense = 0.0
        for month_label, income, expense in monthly_rows:
            total_income += income
            total_expense += expense
            summary_ws.append([month_label, f"{income:.2f}", f"{expense:.2f}"])
        summary_ws.append(["TOTAL", f"{total_income:.2f}", f"{total_expense:.2f}"])

    try:
        groups = report.grouped_by_category()
    except Exception:
        groups = {}

    bycat_ws = wb.create_sheet(title="By Category", index=1)
    for category, subreport in sorted(groups.items(), key=lambda x: x[0] or ""):
        bycat_ws.append([f"Category: {category}"])
        bycat_ws.append(["Date", "Type", "Amount (KZT)"])
        records_total = 0.0
        for r in sorted(
            subreport.records(),
            key=lambda rr: (0, rr.date) if isinstance(rr.date, dt_date) else (1, dt_date.max),
        ):
            typ = record_type_name(r)
            if typ == "income":
                r_type = "Income"
            elif typ == "mandatory_expense":
                r_type = "Mandatory Expense"
            else:
                r_type = "Expense"
            amt = getattr(r, "amount", 0.0)
            records_total += (
                getattr(r, "amount", 0.0) if getattr(r, "amount", None) is not None else 0.0
            )
            display_date = getattr(r, "date", "")
            if isinstance(display_date, dt_date):
                display_date = display_date.isoformat()
            bycat_ws.append([display_date, r_type, f"{abs(amt):.2f}"])
        bycat_ws.append(["SUBTOTAL", "", f"{abs(records_total):.2f}"])
        bycat_ws.append([""])

    os.makedirs(os.path.dirname(filepath), exist_ok=True) if os.path.dirname(filepath) else None
    wb.save(filepath)
    try:
        wb.close()
    except Exception:
        pass
    gc.collect()


def report_from_xlsx(filepath: str) -> Report:
    records, initial_balance, _ = import_records_from_xlsx(filepath, ImportPolicy.LEGACY)
    return Report(records, initial_balance)


def export_records_to_xlsx(
    records: list[Record],
    filepath: str,
    initial_balance: float = 0.0,
    *,
    transfers=None,
) -> None:
    del initial_balance  # legacy argument, kept for compatibility

    transfer_map = {transfer.id: transfer for transfer in (transfers or [])}

    wb = Workbook()
    ws = wb.active
    if ws is not None:
        ws.title = "Data"
        ws.append(DATA_HEADERS)

    for record in records:
        if record.transfer_id is not None:
            continue
        payload = [
            record.date.isoformat() if isinstance(record.date, dt_date) else record.date,
            record_type_name(record),
            int(getattr(record, "wallet_id", 1)),
            record.category,
            record.amount_original,
            record.currency,
            record.rate_at_operation,
            record.amount_kzt,
            str(getattr(record, "description", "") or ""),
            getattr(record, "period", "") if isinstance(record, MandatoryExpenseRecord) else "",
            "",
            "",
            "",
        ]
        if ws is not None:
            ws.append(payload)

    for transfer_id in sorted(transfer_map):
        transfer = transfer_map[transfer_id]
        if ws is not None:
            ws.append(
                [
                    transfer.date.isoformat()
                    if isinstance(transfer.date, dt_date)
                    else transfer.date,
                    "transfer",
                    "",
                    "Transfer",
                    transfer.amount_original,
                    transfer.currency,
                    transfer.rate_at_operation,
                    transfer.amount_kzt,
                    transfer.description,
                    "",
                    transfer.id,
                    transfer.from_wallet_id,
                    transfer.to_wallet_id,
                ]
            )

    os.makedirs(os.path.dirname(filepath), exist_ok=True) if os.path.dirname(filepath) else None
    wb.save(filepath)
    try:
        wb.close()
    except Exception:
        pass
    gc.collect()


def import_records_from_xlsx(
    filepath: str,
    policy: ImportPolicy = ImportPolicy.FULL_BACKUP,
    currency_service=None,
    wallet_ids: set[int] | None = None,
) -> tuple[list[Record], float, ImportSummary]:
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"XLSX file not found: {filepath}")

    wb = load_workbook(filepath, data_only=True)
    try:
        if not wb.worksheets:
            return [], 0.0, (0, 0, [])

        ws = wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return [], 0.0, (0, 0, [])

        header_row_index = 0
        first_row = rows[0] if rows else ()
        if first_row and _safe_str(first_row[0]).strip().startswith("Transaction statement"):
            header_row_index = 1
        if len(rows) <= header_row_index:
            return [], 0.0, (0, 0, [])

        headers = [norm_key(_safe_str(h)) for h in rows[header_row_index]]
        records: list[Record] = []
        initial_balance = 0.0
        errors: list[str] = []
        skipped = 0
        imported = 0

        transfers = {}
        next_transfer_id = 1

        is_report_xlsx = {"date", "type", "category", "amount_(kzt)"}.issubset(set(headers))
        get_rate = None
        if policy == ImportPolicy.CURRENT_RATE:
            get_rate = _resolve_get_rate(currency_service)

        data_rows = rows[header_row_index + 1 :]
        for idx, row in enumerate(data_rows, start=header_row_index + 2):
            if not row or all(cell is None for cell in row):
                continue

            raw = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}

            if is_report_xlsx:
                date_value = _safe_str(raw.get("date", "")).strip()
                if date_value.upper() in {"SUBTOTAL", "FINAL_BALANCE", "FINAL BALANCE"}:
                    continue
                if date_value == "" and norm_key(_safe_str(raw.get("type", "")).strip()) in {
                    "initial_balance",
                    "opening_balance",
                }:
                    raw["type"] = "initial_balance"
                    raw["amount_original"] = raw.get("amount_(kzt)")
                else:
                    raw["amount"] = raw.get("amount_(kzt)")

            row_type = safe_type(_safe_str(raw.get("type", "")).lower())
            if row_type == "transfer":
                row_lc = {norm_key(str(k)): str(v) if v is not None else "" for k, v in raw.items()}
                parsed_records, transfer, next_transfer_id, error = _parse_transfer_row(
                    row_lc=row_lc,
                    row_label=f"row {idx}",
                    policy=policy,
                    get_rate=get_rate,
                    next_transfer_id=next_transfer_id,
                    wallet_ids=wallet_ids,
                )
                if error:
                    skipped += 1
                    errors.append(error)
                    logger.warning("XLSX import skipped %s", error)
                    continue
                if transfer is not None and parsed_records is not None:
                    if transfer.id in transfers:
                        skipped += 1
                        errors.append(f"row {idx}: duplicate transfer_id #{transfer.id}")
                        continue
                    transfers[transfer.id] = transfer
                    records.extend(parsed_records)
                    imported += 1
                continue

            record, parsed_balance, error = parse_import_row(
                raw,
                row_label=f"row {idx}",
                policy=policy,
                get_rate=get_rate,
                mandatory_only=False,
            )
            if error:
                skipped += 1
                errors.append(error)
                logger.warning("XLSX import skipped %s", error)
                continue
            if parsed_balance is not None:
                initial_balance = parsed_balance
                continue
            if record is None:
                continue
            if wallet_ids is not None and record.wallet_id not in wallet_ids:
                skipped += 1
                errors.append(f"row {idx}: wallet not found ({record.wallet_id})")
                continue
            imported += 1
            records.append(record)
            if record.transfer_id is not None:
                next_transfer_id = max(next_transfer_id, int(record.transfer_id) + 1)

        _restore_missing_transfers(records, transfers)
        integrity_errors = _validate_transfer_integrity(records, transfers, wallet_ids)
        if integrity_errors:
            skipped += len(integrity_errors)
            errors.extend(integrity_errors)

        logger.info(
            "XLSX import completed: imported=%s skipped=%s file=%s",
            imported,
            skipped,
            filepath,
        )
        return records, initial_balance, (imported, skipped, errors)
    finally:
        try:
            wb.close()
        except Exception:
            pass
        gc.collect()


def export_mandatory_expenses_to_xlsx(
    expenses: list[MandatoryExpenseRecord], filepath: str
) -> None:
    wb = Workbook()
    ws = wb.active
    if ws is not None:
        ws.title = "Mandatory"
        ws.append(MANDATORY_HEADERS)

    for e in expenses:
        if ws is not None:
            ws.append(
                [
                    e.date.isoformat() if isinstance(e.date, dt_date) else e.date,
                    "mandatory_expense",
                    e.category,
                    e.amount_original,
                    e.currency,
                    e.rate_at_operation,
                    e.amount_kzt,
                    e.description,
                    e.period,
                ]
            )

    os.makedirs(os.path.dirname(filepath), exist_ok=True) if os.path.dirname(filepath) else None
    wb.save(filepath)
    try:
        wb.close()
    except Exception:
        pass
    gc.collect()


def import_mandatory_expenses_from_xlsx(
    filepath: str,
    policy: ImportPolicy = ImportPolicy.FULL_BACKUP,
    currency_service=None,
) -> tuple[list[MandatoryExpenseRecord], ImportSummary]:
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"XLSX file not found: {filepath}")

    wb = load_workbook(filepath, data_only=True)
    try:
        if not wb.worksheets:
            return [], (0, 0, [])

        ws = wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return [], (0, 0, [])

        headers = [norm_key(_safe_str(h)) for h in rows[0]]
        expenses: list[MandatoryExpenseRecord] = []
        errors: list[str] = []
        skipped = 0
        imported = 0

        get_rate = None
        if policy == ImportPolicy.CURRENT_RATE:
            get_rate = _resolve_get_rate(currency_service)

        for idx, row in enumerate(rows[1:], start=2):
            if not row or all(cell is None for cell in row):
                continue
            raw = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}

            record, _, error = parse_import_row(
                raw,
                row_label=f"row {idx}",
                policy=policy,
                get_rate=get_rate,
                mandatory_only=True,
            )
            if error:
                skipped += 1
                errors.append(error)
                logger.warning("Mandatory XLSX import skipped %s", error)
                continue
            if isinstance(record, MandatoryExpenseRecord):
                imported += 1
                expenses.append(record)

        logger.info(
            "Mandatory XLSX import completed: imported=%s skipped=%s file=%s",
            imported,
            skipped,
            filepath,
        )
        return expenses, (imported, skipped, errors)
    finally:
        try:
            wb.close()
        except Exception:
            pass
        gc.collect()
