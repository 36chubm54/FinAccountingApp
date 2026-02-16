import tempfile
import os

from openpyxl import load_workbook

from domain.records import (
    IncomeRecord,
    ExpenseRecord,
    MandatoryExpenseRecord,
)
from domain.reports import Report

from utils.excel_utils import (
    report_to_xlsx,
    report_from_xlsx,
    export_mandatory_expenses_to_xlsx,
    import_mandatory_expenses_from_xlsx,
)

from utils.csv_utils import (
    export_mandatory_expenses_to_csv,
    import_mandatory_expenses_from_csv,
)


def test_report_xlsx_roundtrip():
    records = [
        IncomeRecord(date="2025-01-01", _amount_init=100.0, category="Salary"),
        ExpenseRecord(date="2025-01-02", _amount_init=30.0, category="Food"),
    ]
    report = Report(records, initial_balance=50.0)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name
    try:
        report_to_xlsx(report, tmp_path)
        wb = load_workbook(tmp_path, data_only=True)
        try:
            report_ws = wb["Report"]
            assert report_ws.cell(3, 2).value == "Initial balance"
            assert "Yearly Report" in wb.sheetnames
            summary_ws = wb["Yearly Report"]
            assert summary_ws.cell(1, 1).value == "Month (2025)"
        finally:
            wb.close()
        imported = report_from_xlsx(tmp_path)
        assert len(imported.records()) == 2
        assert abs(imported._initial_balance - 50.0) < 1e-6
        assert abs(imported.total() - report.total()) < 1e-6
    finally:
        os.unlink(tmp_path)


def test_report_xlsx_uses_opening_balance_label_for_filtered_report():
    records = [
        IncomeRecord(date="2024-12-31", _amount_init=20.0, category="Old"),
        IncomeRecord(date="2025-01-01", _amount_init=100.0, category="Salary"),
        ExpenseRecord(date="2025-01-02", _amount_init=30.0, category="Food"),
    ]
    report = Report(records, initial_balance=50.0).filter_by_period("2025")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name
    try:
        report_to_xlsx(report, tmp_path)
        wb = load_workbook(tmp_path, data_only=True)
        try:
            ws = wb["Report"]
            assert ws.cell(3, 2).value == "Opening balance as of 2025-01-01"
            assert ws.cell(3, 4).value == "70.00"
        finally:
            wb.close()
    finally:
        os.unlink(tmp_path)


def test_mandatory_xlsx_roundtrip():
    expenses = [
        MandatoryExpenseRecord(
            date="", _amount_init=10.0, category="Sub", description="d1", period="monthly"
        ),
        MandatoryExpenseRecord(
            date="", _amount_init=20.5, category="Svc", description="d2", period="yearly"
        ),
    ]

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name
    try:
        export_mandatory_expenses_to_xlsx(expenses, tmp_path)
        imported, _ = import_mandatory_expenses_from_xlsx(tmp_path)
        assert len(imported) == 2
        assert imported[0].amount == 10.0
        assert imported[1].period == "yearly"
    finally:
        os.unlink(tmp_path)


def test_mandatory_csv_roundtrip():
    expenses = [
        MandatoryExpenseRecord(
            date="", _amount_init=5.0, category="A", description="x", period="daily"
        ),
    ]
    with tempfile.NamedTemporaryFile(
        delete=False, suffix=".csv", mode="w", newline=""
    ) as tmp:
        tmp_path = tmp.name
    try:
        export_mandatory_expenses_to_csv(expenses, tmp_path)
        imported, _ = import_mandatory_expenses_from_csv(tmp_path)
        assert len(imported) == 1
        assert imported[0].amount == 5.0
        assert imported[0].period == "daily"
    finally:
        os.unlink(tmp_path)
